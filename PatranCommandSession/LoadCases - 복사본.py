import os
import openpyxl as oxl
import sys
from . import PatranCommand
from . import p3Utilities
#from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QAction, QFileDialog
#from PyQt5.QtGui import QIcon


#fname = QFileDialog.getOpenFileName(None, 'Open file', './')

def LoadCase(Inputfile):

    wb = oxl.load_workbook(Inputfile, data_only=True)

    fname = os.path.splitext(Inputfile)[0]

    sht = wb["Input"]

    Action = sht['E1'].value

    loadcase = {}
    lc_name = []
    # lbc_count = []

    iRow = 3
    idx = 0
    while sht.cell(iRow, 2).value != None:
        lc_name.append(sht.cell(iRow, 2).value)

        lbcs = []
        lbc_fac = []
        
        j = 0
        while sht.cell(iRow, 3 + j).value != None:
            lbcs.append(sht.cell(iRow, 3 + j).value)
            lbc_fac.append(float(sht.cell(iRow+1, 3 + j).value))
            
            j += 1

        print('Row : %d'%iRow)
        
        loadcase[lc_name[idx]] = lbcs

        loadcase[lc_name[idx] + '.fac'] = lbc_fac

        idx += 1
        iRow += 2
        
    with open(fname + '.ses', 'w') as f:
        for idx in range(len(lc_name)):
            
            zerof =[0]*len(loadcase[lc_name[idx]])
            
            Session = PatranCommand.create_loadcase(Action, lc_name[idx], loadcase[lc_name[idx]], zerof, loadcase[lc_name[idx] + '.fac'] )
            # 'loadcase_create2( "%s", "Static", "", 1., %s , %s, %s, "", 0., TRUE )\n'%(lc_name[idx], loadcase[lc_name[idx]], zerof, loadcase[lc_name[idx] + '.fac'])

            Session = Session.replace("'","\"")
            Session = p3Utilities.line_breaking(Session)
            '''
            icol = 120
            while Session != "":
                if len(Session) > icol :
                    f.write(Session[:icol] + "@\n")
                else:
                    f.write(Session[:icol] + "\n")    
                
                Session = Session[icol:]
            '''
            f.write(Session)
            f.write('dump %d\n'%(idx+1))

            #print(loadcase[lc_name[idx]])

    f.close()