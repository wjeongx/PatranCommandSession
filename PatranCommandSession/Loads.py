from ast import Not
import os
import openpyxl as oxl
import sys
from . import p3Utilities as UTL
from . import PatranCommand

def load_gen(Inputfile):

    wb = oxl.load_workbook(Inputfile, data_only=True)

    fname = os.path.splitext(Inputfile)[0]

    #wb = oxl.load_workbook('CID Distributed load Input.xlsx', data_only=True)

    sht = wb["Input"]

    Action = sht['F1'].value
    #SF = sht['I1'].value
    #count = sht['E1'].value

    LType = []
    AppType = []
    EntType = []
    lc_name =[]
    con_name =[]
    coord_no = []
    App_Reg =[]
    SF = []
    F = []

    #lc_name =[]
    #con_name =[]
    #coord_no = []
    #App_Reg =[]
    #SL = []
    #EL = []
    #SLF =[]
    #ELF=[]

    iRow = 4

    #for idx in range(count):
    idx = 0

    while sht.cell(iRow, 1).value != None:
        F.append([])
        LType.append(sht.cell(iRow, 1).value)
        AppType.append(sht.cell(iRow, 2).value)
        EntType.append(sht.cell(iRow, 3).value)
        lc_name.append(sht.cell(iRow, 4).value)
        con_name.append(sht.cell(iRow, 5).value)
        coord_no.append(sht.cell(iRow, 6).value)
        App_Reg.append(sht.cell(iRow, 7).value)
        SF.append(sht.cell(iRow, 8).value)

        #start = 9
        for i in range(4):
            start = 9 + i*3 
            cval = [0]*3

            for icnt in range(3):
                cval[icnt] = sht.cell(iRow, start + icnt).value
                if cval[icnt] == None or cval[icnt] == 0:
                    cval[icnt] = ""
                
                print(cval)
                print(LType[idx])

                sval = str(cval)
                sval = sval.replace("''","")

                if sval == "["", "", ""]":
                    sval = ""
                else:
                    sval = sval.replace("[", "<")
                    sval = sval.replace("]",">")
                    

            if sval.find('f:') == -1:
                F[idx].append(sval)
            else:
                F[idx].append(cval[0])
                        
        idx += 1
        iRow += 1

    with open(fname + '.ses', 'w') as f:
        icol = 100
        iRow += 1
        while sht.cell(iRow, 1).value != None:
            xLine = sht.cell(iRow, 1).value
            while xLine != "":
                if len(xLine) > icol :
                    f.write(xLine[:icol] + "@\n")
                else:
                    f.write(xLine[:icol] + "\n")    
                
                xLine = xLine[icol:]
            
            iRow += 1

        for idx in range(len(lc_name)):
            if App_Reg[idx][:2] =="A:":
                App_Reg[idx] = App_Reg[idx][2:]
            elif App_Reg[idx][:2] =="G:":
                App_Reg[idx] = App_Reg[idx][2:]
                f.write('string grp_members[virtual]\n')
                f.write('uil_group_members_get ("%s", grp_members)\n'%App_Reg[idx])
                f.write('string %s[virtual](1)\n'%App_Reg[idx])
                f.write('integer len\n')
                f.write('len = str_length(grp_members)\n')
                f.write('SYS_ALLOCATE_STRING(%s, len+1)\n'%App_Reg[idx])
                f.write('%s(1) = grp_members\n'%App_Reg[idx])
                f.write('dump %s\n'%App_Reg[idx])
            else:
                App_Reg[idx] = '["' + App_Reg[idx] + '"]'

            #f.write('app_list(1) = grp_members\n')
    #        if Action == "Create":
    #            Session =  'loadsbcs_create2( "%s", "CID Distributed Load", "Element Uniform", "2D", "Static", %s , "FEM", "%s", "%s", @\n \
    #            ["<%s,%s,%s>", "<%s,%s,%s>"], ["", ""])\n'%(lc_name[idx]+con_name[idx],App_Reg[idx], coord_no[idx], SF[idx], F[idx][0], F[idx][1], F[idx][1],F[idx][0],F[idx][1],F[idx][1])
                
    #        elif Action == "Modify":
    #            old_name = lc_name[idx]+con_name[idx]
    #            new_name = old_name
    #            Session =  'loadsbcs_modify2( "%s", "%s", "CID Distributed Load", "Element Uniform", "2D", "Static", %s , "FEM", "%s", "%s", @\n \
    #            ["<%s,%s,%s>", "<%s,%s,%s>"], ["", ""] )\n' %(old_name, new_name, App_Reg[idx], coord_no[idx], SF[idx], F[idx][0], F[idx][1], F[idx][1],F[idx][0],F[idx][1],F[idx][1])

            session = PatranCommand.create_load(Action, LType[idx], AppType[idx], EntType[idx], lc_name[idx]+con_name[idx], App_Reg[idx], coord_no[idx], SF[idx], F[idx])

            session = session.replace('None','')
            session = UTL.line_breaking(session)        
            f.write(session + "\n")

    f.close()
    

