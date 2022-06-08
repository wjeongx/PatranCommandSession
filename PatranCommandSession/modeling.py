import os
import openpyxl as oxl
from . import PatranCommand as pcs
def joint_modeling(InputFile):
    wb = oxl.load_workbook(InputFile, data_only=True)
    
    sht = wb["Input"]

    fname = os.path.splitext(InputFile)[0]

    Entity_Type = sht['D1'].value

    id = []
    xyzs =[]
    coord_no = []

    iRow = 3
    # for i in range(count):
    while sht.cell(iRow, 1).value != None:    
        id.append(sht.cell(iRow, 1).value)
        cval = [0]*3
        for icnt in range(3):
            cval[icnt] = sht.cell(iRow, 2 + icnt).value
            if cval[icnt] == None or cval[icnt] == 0:
                cval[icnt] = ""
        
        xyzs.append(cval)

        coord_no.append(sht.cell(iRow, 5).value)

        iRow += 1

    with open(fname + '.ses', 'w') as f:
        for idx in range(len(id)):
            if Entity_Type == "Node":
                session = pcs.create_nodes(id[idx], xyzs[idx], coord_no[idx])
            elif Entity_Type == "Point":
                session = pcs.create_points(id[idx], xyzs[idx], coord_no[idx])
        
            f.write(session)

    f.close()
    

