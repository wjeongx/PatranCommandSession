import os
import openpyxl as oxl
import sys
#import p3Utilities as UTL
from . import p3Utilities
from . import PatranCommand

#from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QAction, QFileDialog
#from PyQt5.QtGui import QIcon

def FEMField(Inputfile):
        
    wb = oxl.load_workbook(Inputfile, data_only=True)

    fname = os.path.splitext(Inputfile)[0]

    sht = wb["Input"]

    count =sht['C1'].value
    Action = sht['A3'].value
    Method = sht['B3'].value
    FieldDef = sht['C3'].value
    FieldTyp = sht['D3'].value
    EntyType = sht['E3'].value

    count = sht['C1'].value

    iRow = 4
    with open(fname+'.ses', 'w') as f:
        for idx in range(count):
            field_name = sht.cell(iRow,1).value
            cnt = sht.cell(iRow,2).value
            iRow += 1
            fld = []
            Entity = []
            for i in range(cnt):
                Entity.append(EntyType + " " + str(sht.cell(iRow, 3).value))
                x = sht.cell(iRow, 4).value
                y = sht.cell(iRow, 5).value
                z = sht.cell(iRow, 6).value
                flds = str([x,y,z]).replace('[', '<')
                flds = flds.replace(']', '>')
                fld.append(flds)
                print(fld[i])

                iRow += 1

            Session = PatranCommand.create_fem_field(Action, field_name, EntyType, FieldTyp, cnt, Entity, fld)
            # 'fields_create_dfem( "%s", "%s", "%s", %d, %s, %s)\n'%(field_name, EntyType, FieldTyp, cnt, Entity, fld)
            # fields_create     ( "Pint1", "Spatial", 1, "Scalar", "Real", "Coord 0", "", "Table", 1, "", "", "Z", "", "", "", FALSE, [0., 40560.], [0.], [0.], [[[0.40799999]][[0.]]] )
            # fields_create_dfem( "cbea2", "Element",    "Vector", 3, ["Elem 42523", "Elem 42524", "Elem 42525"], ["<0., 37.673698, -2.3E-015>", "<0., 37.673698, -2.3E-015>", "<0., 37.673698, -2.3E-015>"] )
            Session = Session.replace("'",'"')
            Session = Session.replace("None",'')

            Session = p3Utilities.line_breaking(Session)

            f.write(Session)

    f.close()

        
#                fields_create_dfem( "test", "Node", "Vector", 2, ["Node 27303", "Node 26437"], ["<1., 0., 0.>", "<1., 0., 0.>"] )