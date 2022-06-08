import os
import openpyxl as oxl
import sys
from . import PatranCommand
from . import p3Utilities
#from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QAction, QFileDialog
#from PyQt5.QtGui import QIcon

def result_combine(Inputfile):
    wb = oxl.load_workbook(Inputfile, data_only=True)

    fname = os.path.splitext(Inputfile)[0]

    sht = wb["Input"]

    lc1 = []
    sc1 = []
    lc2 = []
    sc2 = []
    lc3 = []
    sc3 = []
    irow = 4
    count = sht['C1'].value

    for idx in range(count):
        lc1.append(sht.cell(irow, 2).value)
        sc1.append(sht.cell(irow, 3).value)
        lc2.append(sht.cell(irow, 4).value)
        sc2.append(sht.cell(irow, 5).value)
        lc3.append(sht.cell(irow, 6).value)
        sc3.append(sht.cell(irow, 7).value)
        irow += 1

    with open(fname + '.ses', 'w') as f:
        
        for idx in range(len(lc1)):
            f.write('db_drop_res_index( )\n')
            f.write('INTEGER res_create_demo_lcid\n')
            f.write('res_db_create_loadcase_c( "%s", 1, "Assign Results To A Load Case", res_create_demo_lcid )\n'%lc3[idx])
            f.write('INTEGER res_create_demo_scid\n')
            f.write('INTEGER res_create_demo_rcid\n')
            f.write('dump res_create_demo_lcid\n')
            f.write('res_db_create_subcase_c( res_create_demo_lcid, "Combine Subcase", res_create_demo_scid, res_create_demo_rcid )\n')
            f.write('dump res_create_demo_rcid\n')
            
            f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
            f.write('"Stress Tensor", "", "At Z1","", "", "", "", "", "", 0. )\n')
            f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
            f.write('["%s"], ["Stress Tensor"], [""], ["At Z1", "At Z1"] )\n' %sc2[idx])
            f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
            f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Z1",  @\n'%(lc3[idx], sc3[idx]))
            f.write('"Stress Tensor", "" )\n')              
    
            f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
            f.write('"Stress Tensor", "", "At Z2","", "", "", "", "", "", 0. )\n')
            f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
            f.write('["%s"], ["Stress Tensor"], [""], ["At Z2", "At Z2"] )\n' %sc2[idx])
            f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
            f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Z2",  @\n'%(lc3[idx], sc3[idx]))
            f.write('"Stress Tensor", "" )\n')              
            f.write('db_post_results_load( )\n\n')

            f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
            f.write('"Stress Tensor", "", "At Center","", "", "", "", "", "", 0. )\n')
            f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
            f.write('["%s"], ["Stress Tensor"], [""], ["At Center", "At Center"] )\n' %sc2[idx])
            f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
            f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Center",  @\n'%(lc3[idx], sc3[idx]))
            f.write('"Stress Tensor", "" )\n')              
            f.write('db_post_results_load( )\n\n')

    f.close()

def result_sum(Inputfile):

    wb = oxl.load_workbook(Inputfile, data_only=True)

    fname = os.path.splitext(Inputfile)[0]

    sht = wb["Input"]

    lc1 = []
    sc1 = []
    lc2 = []
    sc2 = []
    lc3 = []
    sc3 = []
    irow = 4
    count = sht['C1'].value

    for idx in range(count):
        lc1.append(sht.cell(irow, 2).value)
        sc1.append(sht.cell(irow, 3).value)
        lc2.append(sht.cell(irow, 4).value)
        sc2.append(sht.cell(irow, 5).value)
        lc3.append(sht.cell(irow, 6).value)
        sc3.append(sht.cell(irow, 7).value)
        irow += 1

    with open(fname + '.ses', 'w') as f:
        for idx in range(len(lc1)):
            f.write('db_drop_res_index( )\n')
            f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
            f.write('"Stress Tensor", "", "At Z1","", "Global", "DeriveAverage", "Element", @ \n')
            f.write('"Centroid", "", 0. )\n')
            f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
            f.write('["%s"], ["Stress Tensor"], [""], ["At Z1"] )\n' %sc2[idx])
            f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
            f.write('INTEGER res_create_drv_maxmin_new_lcid\n')
            f.write('res_db_create_loadcase_c( "%s", 1, "Created by Results Derive",  @\n'%lc3[idx])
            f.write('res_create_drv_maxmin_new_lcid)\n')
            f.write('INTEGER res_create_drv_maxmin_new_scid\n')
            f.write('INTEGER res_create_drv_maxmin_new_rcid\n')
            f.write('integer idx\n')
            f.write('dump idx\n')
            f.write('db_get_load_case_id ("%s", idx)\n'%lc3[idx])
            f.write('res_db_create_subcase_c( idx, "%s", res_create_drv_maxmin_new_scid,  @\n'%sc3[idx])
            f.write('res_create_drv_maxmin_new_rcid )\n')
            f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "",  @\n'%(lc3[idx], sc3[idx]))
            f.write('"Stress Tensor", "", 0, [""] )\n')
            f.write('db_post_results_load( )'+'\n\n')

    f.close()




