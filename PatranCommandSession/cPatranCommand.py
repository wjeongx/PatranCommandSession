from tkinter import filedialog
import openpyxl as oxl
import os

class PatranCommand:
    
    def session_breaking(xLine):
        icol = 96
        Session = ""
        while xLine != "":
            print("xline = %s"%xLine)
            if len(xLine) > icol :
                Session += xLine[:icol] + "@\n"
            else:
                Session += xLine[:icol] + "\n"
                
            xLine = xLine[icol:]

        return Session
    
    def create_points(Idx, xyzs, coord):
        patran_command = 'STRING asm_create_grid_xyz_created_ids[VIRTUAL]\n'
        patran_command += 'asm_const_grid_xyz( "%s","%s", "%s",asm_create_grid_xyz_created_ids)\n'%(Idx, xyzs, coord)
    
        return patran_command

    def create_nodes(Idx, xyzs, coord):
        patran_command = 'STRING fem_create_nodes__nodes_created[VIRTUAL]\n'
        patran_command += 'fem_create_nodes_1( "%s","%s",3,"%s","%s",fem_create_nodes__nodes_created)\n'%(coord, coord, Idx, xyzs )
    
        return patran_command
    
    def create_loadcase(action, lcn, lc, zf, lcf):
        if action == "Create":
            patran_command = 'loadcase_create2( "%s", "Static", "", 1., %s , %s, %s, "", 0., TRUE )\n'%(lcn, lc, zf, lcf)
        elif action == "Modify":
            patran_command = 'loadcase_modify2( "%s", "%s", "Static", "", 1., %s , %s, %s, "", 0., TRUE )\n'%(lcn, lcn, lc, zf, lcf)

        return patran_command

    def create_fem_field(action, field_name, EntyType, FieldTyp, cnt, Entity, fld):
        if action == "Create":
            patran_command = 'fields_create_dfem( "%s", "%s", "%s", %d, %s, %s)\n'%(field_name, EntyType, FieldTyp, cnt, Entity, fld)
        elif action == 'Modify':
            patran_command = 'fields_create_dfem( "%s", "%s", "%s", %d, %s, %s)\n'%(field_name, EntyType, FieldTyp, cnt, Entity, fld)
            patran_command = 'fields_modify_dfem( "%s", "%s", "%s", "%s", %d, %s, %s)\n'%(field_name,field_name, EntyType, FieldTyp, cnt, Entity, fld)
            
        patran_command = patran_command.replace("'",'"')
        patran_command = patran_command.replace("None",'')

        return patran_command


    def create_load(action, load_type, app_type, EntType, ln, app_reg, cno, sf, F):

        ''' load_type
        CID Distributed Load , Force, Total Load
            app_type
        'Element Uniform' for CID Distributed Load, Total Load
        'Nodal' for Force 
        '''
        if load_type == "Force":
            loads = '["%s", "%s", "%s", "%s"], ["", "", "", ""]'%(F[0], F[1], F[2], F[3])
        elif load_type == "Pressure":
            loads = '["%s", "%s", "%s"], ["", "", ""]'%(F[0], F[1], F[2])        
        elif load_type == "CID Distributed Load":
            loads = '["%s", "%s"], ["", ""]'%(F[0], F[1])
        elif load_type == "Total Load":
            loads = '["%s", "%s"], ["", ""]'%(F[0], F[1])
        
        if action == 'Create':    
            patran_command = 'loadsbcs_create2( "%s", "%s", "%s", "%s", "Static", %s , "FEM", "%s", "%s",'%(ln, load_type, app_type, EntType, app_reg, cno, sf)
        elif action == 'Modify':
            patran_command = 'loadsbcs_modify2( "%s", "%s", "%s", "%s", "%s", "Static", %s , "FEM", "%s", "%s",'%(ln, ln, load_type, app_type, EntType, app_reg, cno, sf)
        
        patran_command += loads +')'

        return patran_command

    def result_combine(Inputfile):
        wb = oxl.load_workbook(Inputfile, data_only=True)

        fname = os.path.splitext(Inputfile)[0]

        sht = wb["Input"]

        lc1 = []
        sc1 = []
        lc2 = []
        sc2 = []
        lc3 = []
        sc3 = []
        irow = 4
        count = sht['C1'].value

        for idx in range(count):
            lc1.append(sht.cell(irow, 2).value)
            sc1.append(sht.cell(irow, 3).value)
            lc2.append(sht.cell(irow, 4).value)
            sc2.append(sht.cell(irow, 5).value)
            lc3.append(sht.cell(irow, 6).value)
            sc3.append(sht.cell(irow, 7).value)
            irow += 1

        with open(fname + '.ses', 'w') as f:
            
            for idx in range(len(lc1)):
                f.write('db_drop_res_index( )\n')
                f.write('INTEGER res_create_demo_lcid\n')
                f.write('res_db_create_loadcase_c( "%s", 1, "Assign Results To A Load Case", res_create_demo_lcid )\n'%lc3[idx])
                f.write('INTEGER res_create_demo_scid\n')
                f.write('INTEGER res_create_demo_rcid\n')
                f.write('dump res_create_demo_lcid\n')
                f.write('res_db_create_subcase_c( res_create_demo_lcid, "Combine Subcase", res_create_demo_scid, res_create_demo_rcid )\n')
                f.write('dump res_create_demo_rcid\n')
                
                f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
                f.write('"Stress Tensor", "", "At Z1","", "", "", "", "", "", 0. )\n')
                f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
                f.write('["%s"], ["Stress Tensor"], [""], ["At Z1", "At Z1"] )\n' %sc2[idx])
                f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
                f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Z1",  @\n'%(lc3[idx], sc3[idx]))
                f.write('"Stress Tensor", "" )\n')              
        
                f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
                f.write('"Stress Tensor", "", "At Z2","", "", "", "", "", "", 0. )\n')
                f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
                f.write('["%s"], ["Stress Tensor"], [""], ["At Z2", "At Z2"] )\n' %sc2[idx])
                f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
                f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Z2",  @\n'%(lc3[idx], sc3[idx]))
                f.write('"Stress Tensor", "" )\n')              
                f.write('db_post_results_load( )\n\n')

                f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
                f.write('"Stress Tensor", "", "At Center","", "", "", "", "", "", 0. )\n')
                f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
                f.write('["%s"], ["Stress Tensor"], [""], ["At Center", "At Center"] )\n' %sc2[idx])
                f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
                f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "At Center",  @\n'%(lc3[idx], sc3[idx]))
                f.write('"Stress Tensor", "" )\n')              
                f.write('db_post_results_load( )\n\n')

        f.close()

    def result_sum(Inputfile):

        wb = oxl.load_workbook(Inputfile, data_only=True)

        fname = os.path.splitext(Inputfile)[0]

        sht = wb["Input"]

        lc1 = []
        sc1 = []
        lc2 = []
        sc2 = []
        lc3 = []
        sc3 = []
        irow = 4
        count = sht['C1'].value

        for idx in range(count):
            lc1.append(sht.cell(irow, 2).value)
            sc1.append(sht.cell(irow, 3).value)
            lc2.append(sht.cell(irow, 4).value)
            sc2.append(sht.cell(irow, 5).value)
            lc3.append(sht.cell(irow, 6).value)
            sc3.append(sht.cell(irow, 7).value)
            irow += 1

        with open(fname + '.ses', 'w') as f:
            for idx in range(len(lc1)):
                f.write('db_drop_res_index( )\n')
                f.write('res_data_load_dbresult ( 0, "Element", "Tensor", "%s", "%s" , @\n'%(lc1[idx], sc1[idx]))
                f.write('"Stress Tensor", "", "At Z1","", "Global", "DeriveAverage", "Element", @ \n')
                f.write('"Centroid", "", 0. )\n')
                f.write('res_data_dbres_list( 0, "Element", "Tensor", 1, ["%s"], @\n' %lc2[idx])
                f.write('["%s"], ["Stress Tensor"], [""], ["At Z1"] )\n' %sc2[idx])
                f.write('res_data_list_sum( 0, "Element", "Tensor", 2, [1., 1.] )\n')
                f.write('INTEGER res_create_drv_maxmin_new_lcid\n')
                f.write('res_db_create_loadcase_c( "%s", 1, "Created by Results Derive",  @\n'%lc3[idx])
                f.write('res_create_drv_maxmin_new_lcid)\n')
                f.write('INTEGER res_create_drv_maxmin_new_scid\n')
                f.write('INTEGER res_create_drv_maxmin_new_rcid\n')
                f.write('integer idx\n')
                f.write('dump idx\n')
                f.write('db_get_load_case_id ("%s", idx)\n'%lc3[idx])
                f.write('res_db_create_subcase_c( idx, "%s", res_create_drv_maxmin_new_scid,  @\n'%sc3[idx])
                f.write('res_create_drv_maxmin_new_rcid )\n')
                f.write('res_data_save( 0, "Element", "Tensor", "%s", "%s", "",  @\n'%(lc3[idx], sc3[idx]))
                f.write('"Stress Tensor", "", 0, [""] )\n')
                f.write('db_post_results_load( )'+'\n\n')

        f.close()

    def joint_modeling(InputFile):
        wb = oxl.load_workbook(InputFile, data_only=True)
        
        sht = wb["Input"]

        fname = os.path.splitext(InputFile)[0]

        Entity_Type = sht['D1'].value

        id = []
        xyzs =[]
        coord_no = []

        iRow = 3
        # for i in range(count):
        while sht.cell(iRow, 1).value != None:    
            id.append(sht.cell(iRow, 1).value)
            cval = [0]*3
            for icnt in range(3):
                cval[icnt] = sht.cell(iRow, 2 + icnt).value
                if cval[icnt] == None or cval[icnt] == 0:
                    cval[icnt] = ""
            
            xyzs.append(cval)

            coord_no.append(sht.cell(iRow, 5).value)

            iRow += 1

        with open(fname + '.ses', 'w') as f:
            for idx in range(len(id)):
                if Entity_Type == "Node":
                    session = pcs.create_nodes(id[idx], xyzs[idx], coord_no[idx])
                elif Entity_Type == "Point":
                    session = pcs.create_points(id[idx], xyzs[idx], coord_no[idx])
            
                f.write(session)

        f.close()

        
    